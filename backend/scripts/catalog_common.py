import json
import re
import unicodedata
import zipfile
from pathlib import Path
from typing import Dict, Iterable, List, Optional

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None


REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_ROOT = REPO_ROOT / "backend"
OUTPUT_DIR = BACKEND_ROOT / "output"
POSTERS_DIR = OUTPUT_DIR / "posters"
ANDROID_DRAWABLE_DIR = REPO_ROOT / "HKFilmMap" / "app" / "src" / "main" / "res" / "drawable-nodpi"
CONFIG_DIR = BACKEND_ROOT / "config"
SPREADSHEET_CANDIDATES = [
    REPO_ROOT / "hk_movie_locations.xlsx",
    REPO_ROOT / "hk_movie_locations_bilingual_expanded_2000.xlsx",
]
TMDB_CONFIG_LOCAL_PATH = CONFIG_DIR / "tmdb_config.local.json"
GOOGLE_MAPS_CONFIG_LOCAL_PATH = CONFIG_DIR / "google_maps_config.local.json"
LEGACY_JSON_PATH = REPO_ROOT / "HKFilmMap" / "app" / "src" / "main" / "assets" / "movie_locations.json"
SEED_DB_OUTPUT = OUTPUT_DIR / "hkfilmmap_seed.db"
SEED_DB_ASSET = REPO_ROOT / "HKFilmMap" / "app" / "src" / "main" / "assets" / "hkfilmmap_seed.db"
TMDB_CACHE_PATH = OUTPUT_DIR / "tmdb_movie_metadata.json"
POSTER_MANIFEST_PATH = OUTPUT_DIR / "posters" / "poster_manifest.json"
MISSING_POSTERS_LOG = OUTPUT_DIR / "missing_posters.log"
PLACE_OVERRIDES_PATH = CONFIG_DIR / "place_overrides.json"
MOVIE_OVERRIDES_PATH = CONFIG_DIR / "movie_overrides.json"
CLEANING_REPORT_PATH = OUTPUT_DIR / "cleaning_report.json"

DEFAULT_POSTER_ASSET = "poster_placeholder"

LEGACY_POSTER_MAP = {
    "A Better Tomorrow": "poster_tomorrow",
    "C\u2019est La Vie, Mon Ch\u00e9ri": "poster_cestlavie",
    "Chungking Express": "poster_chungking",
    "Comrades: Almost a Love Story": "poster_comrades",
    "Doctor Strange": "poster_strange",
    "Enter the Dragon": "poster_dragon",
    "Ghost in the Shell": "poster_ghost",
    "Infernal Affairs": "poster_infernal",
    "Police Story": "poster_police",
    "Rouge": "poster_rouge",
    "Rush Hour 2": "poster_rush",
    "The Dark Knight": "poster_darkknight",
    "The Man with the Golden Gun": "poster_golden",
    "The World of Suzie Wong": "poster_suzie",
    "Transformers: Age of Extinction": "poster_transformers",
}

MOVIE_TITLE_ZH_KEYS = ["电影名称", "鐢靛奖鍚嶇О"]
MOVIE_TITLE_EN_KEYS = ["Movie Name"]
MOVIE_GENRE_ZH_KEYS = ["电影类型"]
MOVIE_GENRE_EN_KEYS = ["Movie Genre"]
SCENE_DESC_ZH_KEYS = ["典型场景描述", "鍏稿瀷鍦烘櫙鎻忚堪"]
SCENE_DESC_EN_KEYS = ["Typical Scene Description"]
PLACE_NAME_ZH_KEYS = ["具体地点", "鍏蜂綋鍦扮偣"]
PLACE_NAME_EN_KEYS = ["Specific Location"]
ASCII_STOP_WORDS = {"the", "of", "and", "at", "near", "now", "road", "street", "hong", "kong"}
GENERIC_TITLE_PATTERNS = [
    re.compile(r"^hong kong movie compilation\s+\d+\s*-\s*\d+$", re.IGNORECASE),
    re.compile(r"^other films at\b", re.IGNORECASE),
    re.compile(r"^temple street film\s*\d+$", re.IGNORECASE),
    re.compile(r"^johnnie to films$", re.IGNORECASE),
]


def resolve_spreadsheet_path() -> Path:
    for candidate in SPREADSHEET_CANDIDATES:
        if candidate.exists():
            return candidate
    return SPREADSHEET_CANDIDATES[0]


SPREADSHEET_PATH = resolve_spreadsheet_path()


def ensure_directories() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    POSTERS_DIR.mkdir(parents=True, exist_ok=True)
    ANDROID_DRAWABLE_DIR.mkdir(parents=True, exist_ok=True)
    SEED_DB_ASSET.parent.mkdir(parents=True, exist_ok=True)


def normalize_punctuation(text: str) -> str:
    if not text:
        return ""
    return text.replace("\u2019", "'").replace("\u2018", "'").replace("\u2013", "-").replace("\u2014", "-")


def _is_ascii(text: str) -> bool:
    return all(ord(ch) < 128 for ch in text)


def normalize_key(text: Optional[str]) -> str:
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKC", normalize_punctuation(text)).lower()
    normalized = re.sub(r"\(.*?\)", " ", normalized)
    cleaned = []
    for ch in normalized:
        category = unicodedata.category(ch)
        if category[0] in {"L", "N"}:
            cleaned.append(ch)
        else:
            cleaned.append(" ")
    tokens = [token for token in re.split(r"\s+", "".join(cleaned).strip()) if token]
    filtered = [token for token in tokens if not (_is_ascii(token) and token in ASCII_STOP_WORDS)]
    return " ".join(filtered)


def slugify(text: Optional[str]) -> str:
    if not text:
        return "untitled"
    normalized = unicodedata.normalize("NFKD", normalize_punctuation(text))
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized or "untitled"


def normalize_genre(raw: Optional[str]) -> str:
    text = (raw or "").lower()
    if any(token in text for token in ["romance", "love"]):
        return "Romance"
    if any(token in text for token in ["crime", "gangster", "police", "thriller", "spy"]):
        return "Crime"
    if any(token in text for token in ["comedy", "humour", "humor"]):
        return "Comedy"
    if any(token in text for token in ["science fiction", "sci-fi", "cyberpunk", "superhero", "monster"]):
        return "Sci-Fi"
    if any(token in text for token in ["action", "martial", "adventure", "war"]):
        return "Action"
    if "drama" in text:
        return "Drama"
    return "Various"


def read_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8-sig"))


def write_json(path: Path, payload) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_place_overrides() -> Dict[str, Dict[str, str]]:
    return read_json(PLACE_OVERRIDES_PATH, {"title_aliases": {}, "place_aliases": {}})


def load_movie_overrides() -> Dict[str, Dict[str, object]]:
    return read_json(MOVIE_OVERRIDES_PATH, {"title_aliases": {}, "title_zh_overrides": {}, "tmdb_ids": {}, "year_overrides": {}, "ignored_titles": []})


def get_row_value(row: Dict[str, str], keys: Iterable[str]) -> str:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = normalize_punctuation(str(value).strip())
        if text:
            return text
    return ""


def read_spreadsheet_rows() -> List[Dict[str, str]]:
    if not SPREADSHEET_PATH.exists():
        expected = ", ".join(path.name for path in SPREADSHEET_CANDIDATES)
        raise FileNotFoundError(f"Spreadsheet not found. Expected one of: {expected}")

    if load_workbook is not None:
        workbook = load_workbook(SPREADSHEET_PATH, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in rows[0]]
        records = []
        for row in rows[1:]:
            if not any(value not in (None, "") for value in row):
                continue
            values = ["" if value is None else str(value).strip() for value in row]
            records.append(dict(zip(headers, values)))
        return records

    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    import xml.etree.ElementTree as ET
    with zipfile.ZipFile(SPREADSHEET_PATH) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", ns):
                shared_strings.append("".join((node.text or "") for node in si.iterfind(".//a:t", ns)))

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        first_sheet = workbook.find("a:sheets", ns)[0]
        relation_id = first_sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        worksheet = ET.fromstring(archive.read("xl/" + rel_map[relation_id]))

        rows = []
        for row in worksheet.findall(".//a:sheetData/a:row", ns):
            values = []
            for cell in row.findall("a:c", ns):
                cell_type = cell.attrib.get("t")
                raw_value = cell.find("a:v", ns)
                if raw_value is None:
                    values.append("")
                elif cell_type == "s":
                    values.append(shared_strings[int(raw_value.text)])
                else:
                    values.append(raw_value.text or "")
            rows.append(values)
        headers = rows[0]
        return [dict(zip(headers, row)) for row in rows[1:] if any(value for value in row)]


def load_legacy_rows() -> List[Dict[str, object]]:
    return read_json(LEGACY_JSON_PATH, [])


def is_generic_collection_title(title_en: Optional[str], title_zh: Optional[str]) -> bool:
    en = normalize_punctuation(title_en or "").strip()
    zh = normalize_punctuation(title_zh or "").strip()
    en_lower = en.lower()
    if any(pattern.match(en_lower) for pattern in GENERIC_TITLE_PATTERNS):
        return True
    if "集锦" in zh or "荟萃" in zh:
        return True
    if zh.endswith("等"):
        return True
    return False


def split_multi_movie_entry(title_en: str, title_zh: Optional[str]) -> List[Dict[str, Optional[str]]]:
    separators = ["/", "／"]
    en_parts = [title_en]
    zh_parts = [title_zh] if title_zh else []
    for separator in separators:
        if separator in title_en:
            en_parts = [normalize_punctuation(part.strip()) for part in title_en.split(separator) if part.strip()]
            break
    if title_zh:
        for separator in separators:
            if separator in title_zh:
                zh_parts = [normalize_punctuation(part.strip()) for part in title_zh.split(separator) if part.strip()]
                break
    if len(en_parts) <= 1:
        return [{"titleEn": title_en, "titleZh": title_zh or None}]
    if zh_parts and len(zh_parts) == len(en_parts):
        return [{"titleEn": en_parts[index], "titleZh": zh_parts[index]} for index in range(len(en_parts))]
    return [{"titleEn": part, "titleZh": None} for part in en_parts]


def iter_clean_spreadsheet_rows() -> List[Dict[str, object]]:
    cleaned_rows: List[Dict[str, object]] = []
    skipped: List[Dict[str, object]] = []
    movie_overrides = load_movie_overrides()
    ignored_titles = {normalize_punctuation(str(title).strip()) for title in movie_overrides.get("ignored_titles", [])}
    for source_row, row in enumerate(read_spreadsheet_rows(), start=1):
        title_en = get_row_value(row, MOVIE_TITLE_EN_KEYS)
        title_zh = get_row_value(row, MOVIE_TITLE_ZH_KEYS) or None
        if not title_en:
            skipped.append({"sourceRow": source_row, "reason": "missing_title", "titleEn": title_en, "titleZh": title_zh})
            continue
        movie_genre_en = get_row_value(row, MOVIE_GENRE_EN_KEYS)
        movie_genre_zh = get_row_value(row, MOVIE_GENRE_ZH_KEYS) or None
        scene_desc_en = get_row_value(row, SCENE_DESC_EN_KEYS) or None
        scene_desc_zh = get_row_value(row, SCENE_DESC_ZH_KEYS) or None
        place_name_en = get_row_value(row, PLACE_NAME_EN_KEYS)
        place_name_zh = get_row_value(row, PLACE_NAME_ZH_KEYS) or None

        split_titles = split_multi_movie_entry(title_en, title_zh)
        for variant in split_titles:
            variant_en = variant["titleEn"] or ""
            variant_zh = variant.get("titleZh")
            if variant_en in ignored_titles or (variant_zh and variant_zh in ignored_titles):
                skipped.append({
                    "sourceRow": source_row,
                    "reason": "ignored_title",
                    "titleEn": variant_en,
                    "titleZh": variant_zh,
                })
                continue
            if is_generic_collection_title(variant_en, variant_zh):
                skipped.append({
                    "sourceRow": source_row,
                    "reason": "generic_collection_title",
                    "titleEn": variant_en,
                    "titleZh": variant_zh,
                })
                continue
            cleaned_rows.append({
                "sourceRow": source_row,
                "movieTitleEn": variant_en,
                "movieTitleZh": variant_zh,
                "movieGenreEn": movie_genre_en,
                "movieGenreZh": movie_genre_zh,
                "sceneDescriptionEn": scene_desc_en,
                "sceneDescriptionZh": scene_desc_zh,
                "placeNameEn": place_name_en,
                "placeNameZh": place_name_zh,
            })

    report = {
        "rawRowCount": len(read_spreadsheet_rows()),
        "cleanRowCount": len(cleaned_rows),
        "skippedRowCount": len(skipped),
        "skipped": skipped,
    }
    write_json(CLEANING_REPORT_PATH, report)
    return cleaned_rows


def extract_unique_movies() -> List[Dict[str, object]]:
    rows = iter_clean_spreadsheet_rows()
    legacy_rows = load_legacy_rows()
    movie_overrides = load_movie_overrides()
    title_aliases = movie_overrides.get("title_aliases", {})
    year_overrides = movie_overrides.get("year_overrides", {})
    unique: Dict[str, Dict[str, object]] = {}

    for row in rows:
        title_en = normalize_punctuation(str(row.get("movieTitleEn", "")).strip())
        title_zh = normalize_punctuation(str(row.get("movieTitleZh") or "").strip()) or None
        if not title_en:
            continue
        title_en = title_aliases.get(title_en, title_en)
        key = normalize_key(title_en)
        unique.setdefault(key, {
            "titleEn": title_en,
            "titleZh": title_zh,
            "yearHint": year_overrides.get(title_en),
            "source": "excel",
        })

    for row in legacy_rows:
        title_en = normalize_punctuation(str(row.get("movieTitle", "")).strip())
        if is_generic_collection_title(title_en, str(row.get("movieTitleZh") or "")):
            continue
        title_en = title_aliases.get(title_en, title_en)
        key = normalize_key(title_en)
        if not key:
            continue
        entry = unique.setdefault(key, {
            "titleEn": title_en,
            "titleZh": None,
            "yearHint": None,
            "source": "legacy",
        })
        entry["titleZh"] = entry["titleZh"] or row.get("movieTitleZh") or None
        entry["yearHint"] = entry["yearHint"] or row.get("year")

    return [unique[key] for key in sorted(unique.keys())]

